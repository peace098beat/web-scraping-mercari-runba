# PRG1:ライブラリ設定
import sys
import os
import time
import shutil
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import io
from PIL import Image
import openpyxl as px
from openpyxl.styles import Alignment
 
# メインプログラム
def main():
 
    # PRG2:クロール設定
    BASE_URL = 'https://jp.mercari.com'
    CUR_DIR = os.getcwd()
    args = sys.argv
    KEYWORD = args[1]
    for i in range(2,len(args)):
        KEYWORD += '+' + args[i]
 
    # 検索用URL生成
    options = "&brand_id=3272&category_id=96&price_max=100000&price_min=5000"
    URL_INI = BASE_URL + '/search/?keyword=' + KEYWORD + '&status=on_sale' + options
    url = URL_INI
    page_num = 1
 
    # 検索結果格納リスト・画像保存フォルダ作成
    result = []
    # if os.path.isdir('./img') == False:
        # os.mkdir('./img')
    
    # PRG3:スクレイピング実行
    while True:
        try:
            # ChromeでURLに接続
            # options = Options()
            # options.add_argument('--headless')

            options = webdriver.ChromeOptions()
            options.add_argument('--headless')

            print('connectiong to remote browser...')
            browser = webdriver.Remote(
                command_executor='http://localhost:4444/wd/hub',
                desired_capabilities=options.to_capabilities(),
                options=options,
            )

            # browser = webdriver.Chrome(executable_path='C:\chromedriver.exe', options=options)
            browser.get(url)
            print("wait.. ", url)
            time.sleep(5)


            html = browser.page_source.encode('utf-8')
            soup = BeautifulSoup(html, "html.parser")

            items_list = soup.find_all("li", attrs={'data-testid':'item-cell'})
 
            # サムネイル画像取得
            for i, item in enumerate(items_list):
                print("sleep", i, len(items_list))
                # time.sleep(1)

                # img_src = item.find("mer-item-thumbnail").get('src')
                # response = requests.get(img_src)
                # img_fname = img_src.split('?')[0].split('/')[-1]
                # img_bin = io.BytesIO(response.content)
                # pil_img = Image.open(img_bin)
                # img_resize = scale_to_width(pil_img, 200)
                # img_resize.save('./img/' + img_fname)
 
                # 取得結果をリストに保存
                result.append([item.find("mer-item-thumbnail").get('label'),
                                int(item.find("mer-item-thumbnail").get('price')),
                                "img_url",
                                BASE_URL + item.find('a').get('href')])
                
            # 次ページボタン処理
            next_button = soup.find('mer-button',attrs={'data-testid':"pagination-next-button"})
            if next_button:
                page_num += 1
                param = '&page_token=v1%3A' + str(page_num)
                next_url = URL_INI + param
                url = next_url
                next_url = ''
            else:
                break
            break
 
        # エラー発生時の例外処理
        except Exception as e:
            message = "[エラー]"+"type:{0}".format(type(e))+"\n"+"args:{0}".format(e.args)
            print(message)
 
        # Chrome終了処理
        finally:
            browser.close()
            browser.quit() 
 

    # csvファイルへ書き込み
    import csv
    with open(f'mercari-{KEYWORD}.csv', 'w') as file:
        writer = csv.writer(file, lineterminator='\n')
        writer.writerows(result)

    # # PRG4:Excelに結果出力
    # wb = px.Workbook()
    # ws = wb.active
 
    # # 書式設定
    # ws.column_dimensions['A'].width = 25
    # ws.column_dimensions['B'].width = 18
    # ws.column_dimensions['C'].width = 25
    # ws.column_dimensions['D'].width = 40
    # my_alignment=Alignment(vertical='top', wrap_text=True)
 
    # # Excelヘッダー出力
    # headers = ['アイテム名','価格','サムネイル','URL']
    # for i, header in enumerate(headers):
    #     ws.cell(row=1, column=i+1, value=header)
 
    # # 取得結果書き込み
    # for y, row in enumerate(result):
    #     ws.row_dimensions[y+2].height = 160
    #     for x, cell in enumerate(row):
    #         if x == 2:
    #             ...
    #         #     img = px.drawing.image.Image(cell)
    #         #     img.anchor = ws.cell(row= y+2, column= x+1).coordinate
    #         #     ws.add_image(img)
    #         elif x == 3:
    #             ws.cell(row= y+2, column= x+1).hyperlink = cell
    #             ws.cell(row= y+2, column= x+1).alignment = my_alignment
    #         else:
    #             ws.cell(row= y+2, column= x+1).value = cell
    #             ws.cell(row= y+2, column= x+1).alignment = my_alignment
            
    # # Excelファイル保存
    # xlname = './mercari_' + KEYWORD + '.xlsx'
    # wb.save(xlname)
 
    # # 保存した画像の削除
    # shutil.rmtree('./img')
 
    print('--- END ---')
 
# PRG5:アスペクト比固定して画像リサイズ
def scale_to_width(img, width):  
    height = round(img.height * width / img.width)
    return img.resize((width, height))
 
# スクリプトとして実行された場合の処理
if __name__ == '__main__':
    main()